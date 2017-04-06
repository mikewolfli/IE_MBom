#!/usr/bin/env python
# coding:utf-8
"""
  Author:   10256603<mikewolf.li@tkeap.com>
  Purpose:
  Created: 2016/3/23
"""
import tkinter as tk
from tkinter import *
from tkinter import simpledialog
from tkinter import font
from tkinter import scrolledtext
from tkinter import messagebox
from tkinter import filedialog
import datetime
import tkinter.ttk as ttk
from mbom_dataset import *
from openpyxl import Workbook, load_workbook, reader
import openpyxl.writer.excel as excel_xlsx

import threading
import functools
import ctypes
from tkcalendar import *
import os
import sys
from openpyxl import *
from openpyxl import writer
from openpyxl.drawing.image import Image
from openpyxl.styles import Border, Side, Font
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
import xlrd
import xlwt
from xlutils.copy import copy
from decimal import Decimal
import pyrfc
import threading
import base64
from configparser import ConfigParser
import logging

login_info = {'uid': '', 'pwd': '', 'status': False,
              'perm': '000000', 'plant': '2101'}

NAME = '非标物料处理 '
PUBLISH_KEY = ' R '  # R - release , B - Beta , A- Alpha
VERSION = '2.0.5'
'''
exman 程序集成到此版本中，exman终止。
'''
'''
界面权限：
0 - 无权限
1 - 只读权限
9 - 管理员权限

2 - 电气自制
3 - 钣金自制
4 - PSM
5 - CO Run
6 - 曳引机自制
'''

def move_working_days(start, days, ne=True): 
    day = int(days) 
    hour = (days - day)*8    
    i=0
    if ne:
        while not is_working_day(start):
            start = start + datetime.timedelta(days=1)
            
        da = check_in_worktime(start)
            
        while i<day:
            da = da + datetime.timedelta(days=1)
            if is_working_day(da):
                i+=1
                
        if hour > 0.0:        
            da = check_fin_in_worktime(da, hour)
            
    da=check_in_worktime(da, True)
                 
    return da
        
def check_fin_in_worktime(st, hour):
    ho = st.time().hour
    minute = st.time().minute
    sec = st.time().second
    if (ho+hour==17 and minute==0) or ho+hour<17:
        return st
    else:
        i_hou = ho+hour-17
        st = datetime.datetime(st.date().year, st.date().month, st.date().day, 8, 30, 0)
        st = st+ datetime.timedelta(days=1)
        
        while not is_working_day(st):
            st = st+ datetime.timedelta(days=1)
        
        st = st+datetime.timedelta(hours=i_hou, minutes = minute, seconds=sec)
        
        return st
        
           
def check_in_worktime(st, plus=False):
    ho = st.time().hour
    minu = st.time().minute
    
    if ho < 8:
        da = datetime.datetime(st.date().year, st.date().month, st.date().day, 8, 30, 0)
        if plus:
            da= da+datetime.timedelta(hours=2)
    elif ho==8 and minu<30:
        da = datetime.datetime(st.date().year, st.date().month, st.date().day, 8, 30, 0)
        if plus:
            da= da+datetime.timedelta(hours=2)
    elif ho > 17:
        da = datetime.datetime(st.date().year, st.date().month, st.date().day, 8, 30, 0)
        da = da + datetime.timedelta(days=1)
            
        while not is_working_day(da):
            da = da + datetime.timedelta(days=1)
        
        if plus:
            da= da+datetime.timedelta(hours=2)
    else:
        da = st
        
    return da
                       
def is_working_day(da):
    if type(da)==datetime.datetime:
        de = da.date()
    elif type(da)==datetime.date:
        de = da
        
    try:
        r = working_date.get(working_date.date_desc==de)
        return r.is_working
    except working_date.DoesNotExist:
        d = de.weekday()
        if d<5:
            return True
        else:
            return False



class TextHandler(logging.Handler):
    """This class allows you to log to a Tkinter Text or ScrolledText widget"""

    def __init__(self, text):
        # run the regular Handler __init__
        logging.Handler.__init__(self)
        # Store a reference to the Text it will log to
        self.text = text

    def emit(self, record):
        self.formatter = logging.Formatter(
            '%(asctime)s-%(levelname)s : %(message)s')
        msg = self.format(record)

        def append():
            self.text.configure(state='normal')
            self.text.insert(END, msg + "\n")
            self.text.configure(state='disabled')
            # Autoscroll to the bottom
            self.text.yview(END)
        # This is necessary because we can't modify the Text from other threads
        self.text.after(0, append)  # Scroll to the bottom
        
    
def cur_dir():
    # 获取脚本路径
    path = sys.path[0]
    # 判断为脚本文件还是py2exe编译后的文件，如果是脚本文件，则返回的是脚本的目录，
    # 如果是py2exe编译后的文件，则返回的是编译后的文件路径
    if os.path.isdir(path):
        return path
    elif os.path.isfile(path):
        return os.path.dirname(path)


class date_picker(simpledialog.Dialog):
    result = None

    def body(self, master):
        from_label = Label(master, text='from')
        from_label.grid(row=0, column=0, sticky=EW)
        self.from_var = StringVar()
        self.from_input = Entry(
            master, textvariable=self.from_var, state='readonly')
        self.from_input.grid(row=0, column=1, columnspan=2, sticky=EW)
        to_label = Label(master, text='to')
        to_label.grid(row=1, column=0, sticky=EW)
        self.to_var = StringVar()
        self.to_input = Entry(
            master, textvariable=self.to_var, state='readonly')
        self.to_input.grid(row=1, column=1, columnspan=2, sticky=EW)
        self.from_button = Button(master, text='...')
        self.from_button['command'] = self.from_click
        self.from_button.grid(row=0, column=3)
        self.to_button = Button(master, text='...')
        self.to_button['command'] = self.to_click
        self.to_button.grid(row=1, column=3)
        self.from_var.set(strdate)
        self.to_var.set(strdate)
        return self.from_button

    def from_click(self):
        tkCalendar(self, year, month, day, self.from_var)

    def to_click(self):
        tkCalendar(self, year, month, day, self.to_var)

    def validate(self):
        self.from_date = str2date(self.from_var.get())
        self.to_date = str2date(self.to_var.get())
        if self.from_date is None or self.to_date is None:
            messagebox.showerror('提示', '请务必选择一个日期')
            return 0

        if self.from_date > datetime.datetime.today() or self.to_date > datetime.datetime.today():
            messagebox.showerror('提示', '请选择早于今天的日期')
            return 0

        if self.from_date > self.to_date:
            messagebox.showerror('提示', 'from日期请务必小于to日期')
            return 0

        return 1

    def apply(self):
        self.result = {'from': self.from_date, 'to': self.to_date}


def format_system_message(errno):
    """
    Call FormatMessage with a system error number to retrieve
    the descriptive error message.
    """
    # first some flags used by FormatMessageW
    ALLOCATE_BUFFER = 0x100
    ARGUMENT_ARRAY = 0x2000
    FROM_HMODULE = 0x800
    FROM_STRING = 0x400
    FROM_SYSTEM = 0x1000
    IGNORE_INSERTS = 0x200
    # Let FormatMessageW allocate the buffer (we'll free it below)
    # Also, let it know we want a system error message.
    flags = ALLOCATE_BUFFER | FROM_SYSTEM
    source = None
    message_id = errno
    language_id = 0
    #result_buffer = ctypes.wintypes.LPWSTR()
    result_buffer = ctypes.c_wchar_p()
    buffer_size = 0
    arguments = None
    bytes = ctypes.windll.kernel32.FormatMessageW(
        flags,
        source,
        message_id,
        language_id,
        ctypes.byref(result_buffer),
        buffer_size,
        arguments,
    )
    # note the following will ca`use an infinite loop if GetLastError
    #  repeatedly returns an error that cannot be formatted, although
    #  this should not happen.
    # handle_nonzero_success(bytes)
    message = result_buffer.value
    ctypes.windll.kernel32.LocalFree(result_buffer)
    return message


def handle_nonzero_success(result):
    if result == 0:
        value = ctypes.windll.kernel32.GetLastError()
        strerror = format_system_message(value)
        raise(WindowsError(value, strerror))


def treeview_sort_column(tv, col, reverse):
    l = [(tv.set(k, col), k) for k in tv.get_children('')]
    #i = cols.index(col)

    #l.sort(key=lambda t: t[i], reverse=reverse)
    l.sort(reverse=reverse)
    #      ^^^^^^^^^^^^^^^^^^^^^^^

    for index, (val, k) in enumerate(l):
        tv.move(k, '', index)

    tv.heading(col, command=lambda: treeview_sort_column(tv, col, not reverse))


def center(toplevel):
    toplevel.update_idletasks()
    w = toplevel.winfo_screenwidth()
    h = toplevel.winfo_screenheight()
    size = tuple(int(_) for _ in toplevel.geometry().split('+')[0].split('x'))
    x = w / 2 - size[0] / 2
    y = h / 2 - size[1] / 2
    toplevel.geometry("%dx%d+%d+%d" % (size + (x, y)))


class ScrolledTextDlg(simpledialog.Dialog):
    list_title=''
    def __init__(self, title, method=0, parent=None, initialvalue=None):
        if not parent:
            parent = tk._default_root

        self.initialvalue = initialvalue
        self.method = method
        self.list_title = title

        simpledialog.Dialog.__init__(self, parent, title)
        

    def body(self, master):
        list_title = Label(master, text=self.list_title)
        list_title.pack()
        self.textfield = scrolledtext.ScrolledText(master)
        self.textfield.pack()
        self.textfield.bind_all('<Control-v>', self.copy_ev)
        self.textfield.bind('<Control-V>', self.copy_ev)
        self.textfield.bind("<Next>", self.change_line)
        self.textfield.bind("<Alt-L>", self.change_line)

        if self.initialvalue is not None:
            self.textfield.delete('1.0', END)
            self.textfield.insert(END, self.initialvalue)

        return self.textfield

    def change_line(self, event):
        self.textfield.insert(END, '\n')

    def validate(self):
        try:
            result = self.getresult()
        except ValueError:
            messagebox.showwarning(
                "Illegal value",
                self.errormessage + "\nPlease try again",
                parent=self
            )
            return 0

        res_list = result.split('\n')

        res_res = []
        count = 0
        for res in res_list:
            if len(res.rstrip()) == 0:
                continue

            if len(res.rstrip()) != 9 and (self.method == 0 or self.method == 2):
                messagebox.showwarning("Illegal value", '物料号字符串长度为9位')
                return 0

            if len(res.rstrip()) != 14 and (self.method == 1 or self.method==3):
                messagebox.showwarning("Illegal value", 'WBS No字符串长度为14位')
                return 0

            if self.method == 0 or self.method == 2 :
            #if self.method == 0:
                l = list(res.rstrip())
                for i in range(len(l) - 1, -1, -1):
                    if not(48 <= ord(l[i]) <= 57):
                        messagebox.showwarning("Illegal value", '请输入数值')
                        return 0

            count += 1
            
            if self.method == 1 or self.method==3:
                res_res.append(res.rstrip().upper())
            else:
                res_res.append(res.rstrip())

        if count == 0:
            return 0

        if self.method < 2:
            if messagebox.askyesno('是否继续', '执行数据数量: ' + str(count) + ' 条;此操作不可逆，是否继续(YES/NO)?') == NO:
                return 0

        self.result = res_res
        return 1

    def destroy(self):
        self.textfield = None
        simpledialog.Dialog.destroy(self)

    def getresult(self):
        return self.textfield.get('1.0', END)

    def copy_ev(self, event):
        # self.textfield.delete('1.0',END)
        self.textfield.clipboard_get()


def ask_list(title, method=0):
    d = ScrolledTextDlg(title, method)
    return d.result


def value2key(dic, value):
    if not isinstance(dic, dict):
        return None

    for key, val in dic.items():
        if val == value:
            return key

    return None


def date2str(dt_s):
    if not isinstance(dt_s, datetime.datetime):
        return None
    else:
        return dt_s.strftime("%Y-%m-%d")


def datetime2str(dt_s):
    if not isinstance(dt_s, datetime.datetime):
        return None
    else:
        return dt_s.strftime("%Y-%m-%d %H:%M:%S")


def str2date(dt_s):
    if dt_s is None or (len(dt_s) == 0 and isinstance(dt_s, str)):
        return None
    else:
        return datetime.datetime.strptime(dt_s, '%Y-%m-%d')


def str2datetime(dt_s):
    if dt_s is None or (len(dt_s) == 0 and isinstance(dt_s, str)):
        return None
    else:
        return datetime.datetime.strptime(dt_s, "%Y-%m-%d %H:%M:%S")


def none2str(val):
    if not val:
        return ''
    else:
        return val


def get_name(pid):
    if pid == '' or not pid:
        return ''
    try:
        r_name = s_employee.get(s_employee.employee == pid)
        s_name = r_name.name
    except s_employee.DoesNotExist:
        return 'None'

    return s_name


def change_log(table, section, key, old, new):
    q = s_change_log.insert(table_name=table, change_section=section, key_word=str(key), old_value=str(
        old), new_value=str(new), log_on=datetime.datetime.now(), log_by=login_info['uid'])
    q.execute()

# threads=[]
threadLock = threading.Lock()


class refresh_thread(threading.Thread):

    def __init__(self, pane, typ=None):
        threading.Thread.__init__(self)
        self.pane = pane
        self.type = typ

    def run(self):
        threadLock.acquire()
        self.pane.refresh()
        threadLock.release()
        

class his_display(Toplevel):
    def __init__(self, parent, title, case_list, choice ):
        Toplevel.__init__(self, parent)

        self.withdraw()
        if parent.winfo_viewable():
            self.transient(parent)

        if title:
            self.title(title)

        self.parent = parent
        self.grab_set()
        self.geometry('800x600')

        self.create_widgets(choice)

        self.refresh_list(case_list, choice)

        self.protocol("WM_DELETE_WINDOW", self.close_wm)

        if self.parent is not None:
            center(self)

        self.deiconify() # become visible now

        # wait for window to appear on screen before calling grab_set
        self.wait_visibility()

        self.wait_window(self)

    def create_widgets(self, choice):
        if choice==0:
            self.create_proc_tree()
        elif choice==1:
            self.create_nstd_app_tree()
        elif choice==2 or choice==3:
            self.create_nstd_design_tree()
        elif choice==4:
            self.create_nstd_content()
        elif choice==5 or choice==6:
            self.create_nstd_mat_content()
        ysb = ttk.Scrollbar(self, orient='vertical', command=self.proc_list.yview)
        xsb = ttk.Scrollbar(self, orient='horizontal', command=self.proc_list.xview)
        self.proc_list.grid(row=0, column=0, rowspan=6, columnspan=4, sticky='nsew')
        xsb.grid(row=6, column=0, columnspan=4, sticky='ew')
        ysb.grid(row=0, column=4, rowspan=6, sticky='ns')

        self.close_button= Button(self, text='退出')
        self.close_button.grid(row=7, column=2, columnspan=1, sticky='nsew')
        self.close_button['command']=self.close_wm
        self.export_button = Button(self, text='导出')
        self.export_button.grid(row=7, column=1, sticky='nsew')
        self.export_button['command']=self.export_excel
        if choice!=5 and choice!=6:
            self.export_button.grid_forget()
        self.columnconfigure(3, weight=1)
        self.rowconfigure(5, weight=1)

        self.bind('<Escape>', self.close_wm)

    def export_excel(self):
        items = self.proc_list.get_children('')
        if not items:
            return

        file_str=filedialog.asksaveasfilename(title="导出文件", filetypes=[('excel file','.xlsx')])
        if not file_str:
            return

        if not file_str.endswith(".xlsx"):
            file_str+=".xlsx"

        wb=Workbook()
        ws=wb.worksheets[0]
        ws.title='物料清单'
        col_size = len(self.cols_his)
        for i in range(col_size):
            ws.cell(row=1,column=i+1).value=self.tree_head_his[i]
        n=0

        for item in items:
            if self.proc_list.parent(item) !='':
                continue

            for i in range(col_size):
                ws.cell(row=n+2, column=i+1).value = self.proc_list.item(item, 'values')[i]

            n+=1

        if excel_xlsx.save_workbook(workbook=wb, filename=file_str):
            messagebox.showinfo("输出","成功输出!")

    def close_wm(self, event=None):
        self.withdraw()
        self.update_idletasks()
        Toplevel.destroy(self)

    def create_proc_tree(self):
        self.proc_list = ttk.Treeview(self, show='headings', columns=['col1','col2','col3','col4','col5','col6','col7','col8','col9','col10'])
        self.proc_list.heading('col1', text='Unit WBS NO.')
        self.proc_list.heading('col2', text='项目信息')
        self.proc_list.heading('col3', text='梯号')
        self.proc_list.heading('col4', text='流程')
        self.proc_list.heading('col5', text='步骤')
        self.proc_list.heading('col6', text='负责人')
        self.proc_list.heading('col7', text='收到日期')
        self.proc_list.heading('col8', text='完成日期')
        self.proc_list.heading('col9', text='配置完成日期')
        self.proc_list.heading('col10', text='发货期')
        self.proc_list.column('col1', width=100, anchor='w')
        self.proc_list.column('col2', width=100, anchor='w')
        self.proc_list.column('col3', width=50, anchor='w')
        self.proc_list.column('col4', width=100, anchor='w')
        self.proc_list.column('col5', width=100, anchor='w')
        self.proc_list.column('col6', width=100, anchor='w')
        self.proc_list.column('col7', width=100, anchor='w')
        self.proc_list.column('col8', width=100, anchor='w')
        self.proc_list.column('col9', width=100, anchor='w')
        self.proc_list.column('col10', width=100, anchor='w')

    def create_nstd_app_tree(self):
        self.proc_list = ttk.Treeview(self, show='headings', columns=['col1','col2','col3','col4','col5','col6','col7','col8','col9'])
        self.proc_list.heading('col1', text='非标编号')
        self.proc_list.heading('col2', text='项目信息')
        self.proc_list.heading('col3', text='流程')
        self.proc_list.heading('col4', text='步骤')
        self.proc_list.heading('col5', text='负责人')
        self.proc_list.heading('col6', text='收到日期')
        self.proc_list.heading('col7', text='完成日期')
        self.proc_list.heading('col8', text='配置完成日期')
        self.proc_list.heading('col9', text='发货期')
        self.proc_list.column('col1', width=100, anchor='w')
        self.proc_list.column('col2', width=100, anchor='w')
        self.proc_list.column('col3', width=100, anchor='w')
        self.proc_list.column('col4', width=100, anchor='w')
        self.proc_list.column('col5', width=100, anchor='w')
        self.proc_list.column('col6', width=100, anchor='w')
        self.proc_list.column('col7', width=100, anchor='w')
        self.proc_list.column('col8', width=100, anchor='w')
        self.proc_list.column('col9', width=100, anchor='w')

    def create_nstd_design_tree(self):
        self.proc_list = ttk.Treeview(self, show='headings', columns=['col1','col2','col3','col4','col5','col6','col7','col8','col9','col10'])
        self.proc_list.heading('col1', text='非标编号')
        self.proc_list.heading('col2', text='非标物料申请号')
        self.proc_list.heading('col3', text='项目信息')
        self.proc_list.heading('col4', text='流程')
        self.proc_list.heading('col5', text='步骤')
        self.proc_list.heading('col6', text='负责人')
        self.proc_list.heading('col7', text='收到日期')
        self.proc_list.heading('col8', text='完成日期')
        self.proc_list.heading('col9', text='配置完成日期')
        self.proc_list.heading('col10', text='发货期')
        self.proc_list.column('col1', width=100, anchor='w')
        self.proc_list.column('col2', width=100, anchor='w')
        self.proc_list.column('col3', width=100, anchor='w')
        self.proc_list.column('col4', width=100, anchor='w')
        self.proc_list.column('col5', width=100, anchor='w')
        self.proc_list.column('col6', width=100, anchor='w')
        self.proc_list.column('col7', width=100, anchor='w')
        self.proc_list.column('col8', width=100, anchor='w')
        self.proc_list.column('col9', width=100, anchor='w')
        self.proc_list.column('col10', width=100, anchor='w')

    def create_nstd_content(self):
        self.proc_list = ttk.Treeview(self, show='headings', columns=['col1','col2','col3','col4','col5','col6','col7','col8','col9','col10','col11','col12','col13'])
        self.proc_list.heading('col1', text='非标任务编号')
        self.proc_list.heading('col2', text='非标物料申请号')
        self.proc_list.heading('col3', text='项目信息')
        self.proc_list.heading('col4', text='物料需求日')
        self.proc_list.heading('col5', text='图纸需求日')
        self.proc_list.heading('col6', text='非标类别')
        self.proc_list.heading('col7', text='非标原因')
        self.proc_list.heading('col8', text='分值分布')
        self.proc_list.heading('col9', text='非标负责人')
        self.proc_list.heading('col10', text='分任务描述')
        self.proc_list.heading('col11', text='分任务非标工程师')
        self.proc_list.heading('col12', text='分任务状态')
        self.proc_list.heading('col13', text='关联梯')
        self.proc_list.column('col1', width=80, anchor='w')
        self.proc_list.column('col2', width=80, anchor='w')
        self.proc_list.column('col3', width=100, anchor='w')
        self.proc_list.column('col4', width=100, anchor='w')
        self.proc_list.column('col5', width=100, anchor='w')
        self.proc_list.column('col6', width=100, anchor='w')
        self.proc_list.column('col7', width=100, anchor='w')
        self.proc_list.column('col8', width=100, anchor='w')
        self.proc_list.column('col9', width=100, anchor='w')
        self.proc_list.column('col10', width=100, anchor='w')
        self.proc_list.column('col11', width=100, anchor='w')
        self.proc_list.column('col12', width=100, anchor='w')
        self.proc_list.column('col13', width=100, anchor='w')

    def create_nstd_mat_content(self):
        self.cols_his = ['col0','col1','col2','col3','col4','col5','col6','col7','col8','col9','col10','col11','col12','col13','col14','col15','col16','col17','col18','col19','col20','col21']
        self.tree_head_his=['导入日期','非标编号','判断','物料号','物料名称(中)','物料名称(英)','图号','单位','备注','RP','BoxId','申请人','自制完成','负责人','完成日期','价格维护','负责人','完成日期','CO-run','负责人','完成日期','物料要求维护完成日期']
        self.proc_list = ttk.Treeview(self, show='headings',columns=self.cols_his)
        for col in self.cols_his:
            i = self.cols_his.index(col)
            self.proc_list.heading(col, text=self.tree_head_his[i], command=lambda _col=col: treeview_sort_column(self.proc_list, _col, False))

        self.proc_list.column('col0', width=80, anchor='w')
        self.proc_list.column('col1', width=100, anchor='w')
        self.proc_list.column('col2', width=80, anchor='w')
        self.proc_list.column('col3', width=80, anchor='w')
        self.proc_list.column('col4', width=100, anchor='w')
        self.proc_list.column('col5', width=100, anchor='w')
        self.proc_list.column('col6', width=100, anchor='w')
        self.proc_list.column('col7', width=40, anchor='w')
        self.proc_list.column('col8', width=200, anchor='w')
        self.proc_list.column('col9', width=50, anchor='w')
        self.proc_list.column('col10', width=50, anchor='w')
        self.proc_list.column('col11', width=60, anchor='w')
        self.proc_list.column('col12', width=60, anchor='w')
        self.proc_list.column('col13', width=50, anchor='w')
        self.proc_list.column('col14', width=100, anchor='w')
        self.proc_list.column('col15', width=60, anchor='w')
        self.proc_list.column('col16', width=50, anchor='w')
        self.proc_list.column('col17', width=100, anchor='w')
        self.proc_list.column('col18', width=60, anchor='w')
        self.proc_list.column('col19', width=50, anchor='w')
        self.proc_list.column('col20', width=100, anchor='w')
        self.proc_list.column('col21', width=100, anchor='w')

    def     refresh_list(self, cases, choice):
        for item in self.proc_list.get_children():
            self.proc_list.delete(item)

        if choice==0:
            q_res = LProcAct.select(LProcAct, UnitInfo,ProjectInfo, SEmployee.name,WorkflowInfo.workflow_name, ActionInfo.action_name).join(UnitInfo, on=(LProcAct.instance==UnitInfo.wbs_no)).\
                            switch(UnitInfo).join(ProjectInfo, on=(ProjectInfo.project==UnitInfo.project)).switch(LProcAct).join(WorkflowInfo, on=(WorkflowInfo.workflow==LProcAct.workflow)).join(ActionInfo, on=(ActionInfo.action==LProcAct.action)).join(SEmployee, on=(LProcAct.operator==SEmployee.employee)).\
                            where(((LProcAct.workflow=='WF0002')|(LProcAct.workflow=='WF0006'))& (LProcAct.instance.in_(cases))).order_by(UnitInfo.wbs_no.asc(),LProcAct.flow_ser.asc()).naive()
        elif choice==1:
            indexes=[]
            for r in cases:
                index_res = NonstdAppItem.select(NonstdAppItem.index).where(NonstdAppItem.link_list.contains(r))
                if not index_res:
                    continue
                for i in index_res:
                    if i.index not in indexes:
                        indexes.append(i.index)
                    else:
                        continue

            if len(indexes)==0:
                return

            q_res = LProcAct.select(LProcAct, ProjectInfo,WorkflowInfo, ActionInfo, SEmployee.name,NonstdAppHeader.mat_req_date, NonstdAppHeader.drawing_req_date).join(NonstdAppItem, on=(NonstdAppItem.index==LProcAct.instance)).switch(NonstdAppItem).join(NonstdAppHeader, on=(NonstdAppItem.nonstd==NonstdAppHeader.nonstd)).\
                        switch(NonstdAppHeader).join(ProjectInfo, on=(NonstdAppHeader.project==ProjectInfo.project)).switch(LProcAct).join(WorkflowInfo, on=(WorkflowInfo.workflow==LProcAct.workflow)).join(ActionInfo, on=(ActionInfo.action==LProcAct.action)).join(SEmployee, on=(LProcAct.operator==SEmployee.employee)).\
                        where((LProcAct.instance.in_(indexes))&(NonstdAppHeader.status>=0)).order_by(LProcAct.instance.asc(), LProcAct.flow_ser.asc()).naive()
        elif choice==2:
            indexes=[]
            for r in cases:
                index_res = NonstdAppItemInstance.select(NonstdAppItemInstance.index_mat).join(NonstdAppItem, on = (NonstdAppItemInstance.index==NonstdAppItem.index)).where(NonstdAppItem.link_list.contains(r))
                if not index_res:
                    continue
                for i in index_res:
                    if i.index_mat not in indexes:
                        indexes.append(i.index_mat)
                    else:
                        continue
            if len(indexes)==0:
                return

            q_res = LProcAct.select(LProcAct, ProjectInfo,WorkflowInfo, ActionInfo, SEmployee.name, NonstdAppItemInstance.nstd_mat_app,NonstdAppHeader.mat_req_date, NonstdAppHeader.drawing_req_date).\
                        join(NonstdAppItemInstance, on=(NonstdAppItemInstance.index_mat==LProcAct.instance)).switch(NonstdAppItemInstance).join(NonstdAppItem, on=(NonstdAppItemInstance.index==NonstdAppItem.index)).switch(NonstdAppItem).join(NonstdAppHeader, on=(NonstdAppHeader.nonstd==NonstdAppItem.nonstd)).switch(NonstdAppHeader).\
                        join(ProjectInfo, on=(NonstdAppHeader.project==ProjectInfo.project)).switch(LProcAct).join(WorkflowInfo, on=(WorkflowInfo.workflow==LProcAct.workflow)).join(ActionInfo, on=(ActionInfo.action==LProcAct.action)).join(SEmployee, on=(LProcAct.operator==SEmployee.employee)).\
                        where((LProcAct.instance.in_(indexes))&(NonstdAppItemInstance.status>=0)&(LProcAct.workflow=='WF0005')).order_by(LProcAct.instance.asc(), LProcAct.flow_ser.asc()).naive()
        elif choice==3:
            indexes=[]
            for r in cases:
                index_res = NonstdAppItemInstance.select(NonstdAppItemInstance.index_mat).join(NonstdAppItem, on = (NonstdAppItemInstance.index==NonstdAppItem.index)).where(NonstdAppItem.link_list.contains(r))
                if not index_res:
                    continue
                for i in index_res:
                    if i.index_mat not in indexes:
                        indexes.append(i.index_mat)
                    else:
                        continue
            if len(indexes)==0:
                return

            q_res = LProcAct.select(LProcAct,ProjectInfo,WorkflowInfo,ActionInfo,NonstdAppItemInstance.nstd_mat_app,SEmployee.name,NonstdAppHeader.mat_req_date, NonstdAppHeader.drawing_req_date).\
                        join(NonstdAppItemInstance, on=(NonstdAppItemInstance.index_mat==LProcAct.instance)).switch(NonstdAppItemInstance).join(NonstdAppItem, on=(NonstdAppItemInstance.index==NonstdAppItem.index)).switch(NonstdAppItem).join(NonstdAppHeader, on=(NonstdAppHeader.nonstd==NonstdAppItem.nonstd)).switch(NonstdAppHeader).\
                        join(ProjectInfo, on=(NonstdAppHeader.project==ProjectInfo.project)).switch(LProcAct).join(WorkflowInfo, on=(WorkflowInfo.workflow==LProcAct.workflow)).join(ActionInfo, on=(ActionInfo.action==LProcAct.action)).join(SEmployee, on=(LProcAct.operator==SEmployee.employee)).\
                        where((LProcAct.instance.in_(indexes))&(NonstdAppItemInstance.status>=0)&(LProcAct.workflow=='WF0004')).order_by(LProcAct.instance.asc(), LProcAct.flow_ser.asc()).naive()
        elif choice==4:
            indexes=[]
            for r in cases:
                index_res = NonstdAppItemInstance.select(NonstdAppItemInstance.index_mat).join(NonstdAppItem, on = (NonstdAppItemInstance.index==NonstdAppItem.index)).where(NonstdAppItem.link_list.contains(r))
                if not index_res:
                    continue
                for i in index_res:
                    if i.index_mat not in indexes:
                        indexes.append(i.index_mat)
                    else:
                        continue

            if len(indexes)==0:
                return

            q_res = NonstdAppItemInstance.select(NonstdAppItemInstance, NonstdAppItem, NonstdAppHeader.mat_req_date, NonstdAppHeader.drawing_req_date, ProjectInfo, SEmployee.name).join(NonstdAppItem, on=(NonstdAppItemInstance.index==NonstdAppItem.index)).switch(NonstdAppItem).join(NonstdAppHeader, on=(NonstdAppHeader.nonstd==NonstdAppItem.nonstd)).switch(NonstdAppHeader).\
                       join(ProjectInfo, on=(ProjectInfo.project==NonstdAppHeader.project)).switch(NonstdAppItemInstance).join(SEmployee, on=(NonstdAppItemInstance.res_engineer==SEmployee.employee)).\
                       where((NonstdAppItemInstance.index_mat.in_(indexes))&(NonstdAppItemInstance.status>=0)).order_by(NonstdAppItemInstance.index_mat.asc()).naive()
        elif choice==5:
            q_res = nstd_app_head.select(nstd_app_head, nstd_mat_table, nstd_mat_fin).join(nstd_app_link).switch(nstd_app_head).join(nstd_mat_table).switch(nstd_mat_table).join(nstd_mat_fin).\
                        where((nstd_app_link.wbs_no.in_(cases))&(nstd_mat_fin.justify>=0)).naive()
        elif choice==6:
            q_res = nstd_app_head.select(nstd_app_head, nstd_mat_table, nstd_mat_fin).join(nstd_mat_table).switch(nstd_mat_table).join(nstd_mat_fin)\
                .where(((((nstd_mat_fin.justify == 1) | (nstd_mat_fin.justify == 2) | (nstd_mat_fin.justify == 6)) & (nstd_mat_fin.pu_price_fin == True)) | (((nstd_mat_fin.justify == 3) | (nstd_mat_fin.justify == 4) | (nstd_mat_fin.justify == 5)) & (nstd_mat_fin.mbom_fin == True))) & (nstd_mat_fin.co_run_fin == False)).naive()
                
        self.insert_data(q_res, choice)

    def insert_data(self, q_res, choice):
        if not q_res:
            return
        init_str=''
        b_switch=False
        if choice==0:
            for r in q_res:
                item=[]
                if len(init_str)!=0 and init_str!=r.instance:
                    b_switch = not b_switch
                init_str=r.instance
                item.append(r.instance)
                item.append(r.contract+' '+r.project_name)
                item.append(r.lift_no)
                item.append(r.workflow_name)
                item.append(r.action_name)
                item.append(r.name)
                item.append(none2str(datetime2str(r.start_date)))
                b_active = r.is_active
                if not b_active:
                    item.append(none2str(datetime2str(r.finish_date)))
                else:
                    item.append('')
                item.append(none2str(date2str(r.req_configure_finish)))
                item.append(none2str(date2str(r.req_delivery_date)))
                if b_switch:
                    self.proc_list.insert('', END, values=item, tags=('switch',))
                elif not b_active:
                    self.proc_list.insert('', END, values=item, tags=('unactive',))
                else:
                    self.proc_list.insert('', END, values=item, tags=('active',))
        elif choice==1:
            for r in q_res:
                item=[]
                if len(init_str)!=0 and init_str!=r.instance:
                    b_switch = not b_switch
                init_str=r.instance
                item.append(r.instance)
                item.append(r.contract+' '+r.project_name+' '+r.project)
                item.append(r.workflow_name)
                item.append(r.action_name)
                item.append(r.name)
                item.append(none2str(datetime2str(r.start_date)))
                b_active=r.is_active
                if not b_active:
                    item.append(none2str(datetime2str(r.finish_date)))
                else:
                    item.append('')
                item.append(none2str(date2str(r.mat_req_date)))
                item.append(none2str(date2str(r.drawing_req_date)))
                if b_switch:
                    self.proc_list.insert('', END, values=item, tags=('switch',))
                elif not b_active:
                    self.proc_list.insert('', END, values=item, tags=('unactive',))
                else:
                    self.proc_list.insert('', END, values=item, tags=('active',))
        elif choice==2 or choice==3:
            for r in q_res:
                item=[]
                if len(init_str)!=0 and init_str!=r.instance:
                    b_switch = not b_switch
                init_str=r.instance
                item.append(r.instance)
                item.append(r.nstd_mat_app)
                item.append(r.contract+' '+r.project_name+' '+r.project)
                item.append(r.workflow_name)
                item.append(r.action_name)
                item.append(r.name)
                item.append(none2str(datetime2str(r.start_date)))
                b_active=r.is_active
                if not b_active:
                    item.append(none2str(datetime2str(r.finish_date)))
                else:
                    item.append('')
                item.append(none2str(date2str(r.mat_req_date)))
                item.append(none2str(date2str(r.drawing_req_date)))
                if b_switch:
                    self.proc_list.insert('', END, values=item, tags=('switch',))
                elif not b_active:
                    self.proc_list.insert('', END, values=item, tags=('unactive',))
                else:
                    self.proc_list.insert('', END, values=item, tags=('active',))
        elif choice==4:
            for r in q_res:
                item=[]
                item.append(r.index_mat)
                item.append(r.nstd_mat_app)
                item.append(r.contract+' '+r.project_name+' '+r.project)
                item.append(none2str(date2str(r.mat_req_date)))
                item.append(none2str(date2str(r.drawing_req_date)))
                item.append(none2str(r.nonstd_catalog))
                item.append(none2str(r.nonstd_desc))
                item.append(none2str(r.nonstd_value))
                s_person=r.res_person
                try:
                    s_res_person= SEmployee.get(SEmployee.employee==s_person).name
                except SEmployee.DoesNotExist:
                    s_res_person=s_person
                item.append(s_res_person)
                item.append(none2str(r.instance_nstd_desc))
                item.append(none2str(r.name))
                item.append(Status_Types[r.status])
                item.append(none2str(r.link_list))
                self.proc_list.insert('', END, values=item)
        elif choice==5 or choice==6:
            for r in q_res:
                item=[]
                nstd_app_id = r.nstd_app
                #index_mat_id = r.index_mat
                item.append(none2str(date2str(r.modify_on)))
                item.append(nstd_app_id)
                item.append(Justify_Types[r.justify])
                item.append(r.mat_no)
                item.append(r.mat_name_cn)
                item.append(r.mat_name_en)
                item.append(none2str(r.drawing_no))
                item.append(r.mat_unit)
                item.append(none2str(r.comments))
                item.append(none2str(r.rp))
                item.append(none2str(r.box_code_sj))
                app_per = r.app_person
                if app_per is None:
                    app_per = r.mat_app_person
                elif len(app_per)==0:
                    app_per = r.mat_app_person

                item.append(app_per)
                temp = (r.mbom_fin and 'Y' or '')
                item.append(temp)
                if temp.upper()=='Y':
                    item.append(get_name(r.mbom_fin_by))
                    item.append(none2str(r.mbom_fin_on))
                else:
                    item.append('')
                    item.append('')
                    
                temp=(r.pu_price_fin and 'Y' or '')
                item.append(temp)
                if temp.upper()=='Y':
                    item.append(get_name(r.pu_price_fin_by))
                    item.append(none2str(r.pu_price_fin_on))
                else:
                    item.append('')
                    item.append('')
                    
                temp = (r.co_run_fin and 'Y' or '')
                item.append(temp)
                if temp.upper()=='Y':
                    item.append(get_name(r.co_run_fin_by))
                    item.append(none2str(r.co_run_fin_on))
                else:
                    item.append('')
                    item.append('')
                
                if choice==5:   
                    item.append(date2str(r.req_fin_on))
                elif choice==6:
                    item.append(datetime2str(r.req_fin_on))

                if temp.upper()=='Y':
                    self.proc_list.insert('', END, values=item, tags=('unactive',))
                else:
                    self.proc_list.insert('', END, values=item, tags=('active',))

        self.proc_list.tag_configure('unactive', background='lightgrey')
        self.proc_list.tag_configure('active', background='lightpink')
        self.proc_list.tag_configure('switch', background='white')

